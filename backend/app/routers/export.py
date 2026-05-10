from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.client import Client
from app.models.visit import Visit
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
from datetime import datetime

router = APIRouter()

@router.get("/moh512")
def export_moh512(db: Session = Depends(get_db)):
    clients = db.query(Client).all()
    visits = db.query(Visit).order_by(Visit.visit_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MOH 512"

    headers = [
        "A: Visit Date","B: Reg Number","C: First Name","C: Last Name",
        "D: Visit Type","E: First FP User","F: Age","G: Sex",
        "H: Disability","I: Telephone","J: Location",
        "Weight (kg)","Height (cm)","BMI","BP Systolic","BP Diastolic",
        "PDT Done","PDT Result","Pregnancy Ruled Out",
        "MEC Conditions","Primary Method","Qty Dispensed",
        "DMPA Admin","Take-Home Doses",
        "AJ: HIV Counselled","AK: HIV Tested","AL: HIV Status",
        "AI: TB Status","AM: IPV Status",
        "AN: Cervical Method","AO: Cervical Result",
        "Condom Type","Condom Qty",
        "AF: Natural FP","AG: Cycle Beads",
        "AH: PPFP Timing","AP: Referral In","AQ: Referral Out",
        "AR: Remarks","Return Date"
    ]

    # Header styling
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 18

    ws.row_dimensions[1].height = 40

    client_map = {str(c.id): c for c in clients}

    for row_num, v in enumerate(visits, 2):
        c = client_map.get(str(v.client_id))
        if not c:
            continue
        row_data = [
            v.visit_date.strftime('%d/%m/%Y') if v.visit_date else '',
            c.service_registration_number or '',
            c.first_name or '',
            c.last_name or '',
            'New' if v.visit_type == 1 else 'Revisit',
            'Y' if v.first_ever_user else 'N',
            c.age or '',
            c.sex or '',
            c.disability_status or 0,
            c.telephone or '',
            c.location_landmark or '',
            v.weight_kg or '',
            v.height_cm or '',
            v.bmi_calculated or '',
            v.bp_systolic or '',
            v.bp_diastolic or '',
            'Y' if v.pdt_done else 'N',
            v.pdt_result or '',
            'Y' if (v.pregnancy_checklist and any(v.pregnancy_checklist.values())) else 'N',
            ', '.join(v.mec_conditions) if v.mec_conditions else '',
            v.primary_method or '',
            v.quantity_dispensed or '',
            v.dmpa_sc_admin_type or '',
            v.dmpa_sc_take_home_doses or '',
            'Y' if v.hiv_counselled else 'N',
            'Y' if v.hiv_tested else 'N',
            v.hiv_status or '',
            v.tb_status or '',
            v.ipv_rc_status or '',
            v.cervical_screening_method or '',
            v.cervical_screening_result or '',
            v.condoms_client_sex or '',
            v.condoms_qty_dispensed or '',
            'Y' if v.natural_fp_counselled else 'N',
            'Y' if v.cycle_beads_given else 'N',
            v.ppfp_timing or '',
            v.referral_in or '',
            v.referral_out or '',
            v.remarks or '',
            v.return_date.strftime('%d/%m/%Y') if v.return_date else '',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = Alignment(horizontal='left')
            if row_num % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F0F4F8")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"MOH512_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/moh512/report")
def get_moh512_report(
    month: int = None,
    year: int = None,
    db: Session = Depends(get_db)
):
    from sqlalchemy import extract

    query = db.query(Visit)

    if month and year:
        query = query.filter(
            extract('month', Visit.visit_date) == month,
            extract('year', Visit.visit_date) == year
        )
    elif year:
        query = query.filter(extract('year', Visit.visit_date) == year)

    visits = query.order_by(Visit.visit_date.asc()).all()

    clients_map = {}
    result = []

    for v in visits:
        client_id = str(v.client_id)
        if client_id not in clients_map:
            c = db.query(Client).filter(Client.id == v.client_id).first()
            clients_map[client_id] = c
        c = clients_map[client_id]
        if not c:
            continue

        result.append({
            "visit_id": str(v.id),
            "visit_date": v.visit_date.strftime('%d/%m/%Y') if v.visit_date else '',
            "visit_date_iso": v.visit_date.isoformat() if v.visit_date else '',
            "is_anonymous": getattr(v, 'is_anonymous', False) or False,
            # Client info
            "reg_number": c.service_registration_number or '',
            "first_name": c.first_name or '',
            "last_name": c.last_name or '',
            "age": c.age or '',
            "sex": c.sex or '',
            "anon_age_bracket": getattr(v, 'anon_age_bracket', '') or '',
            "anon_sex": getattr(v, 'anon_sex', '') or '',
            "disability_status": c.disability_status or 0,
            "telephone": c.telephone or '',
            "location": c.location_landmark or '',
            # Visit
            "visit_type": v.visit_type or 1,
            "first_ever_user": v.first_ever_user or False,
            "provider_name": v.provider_name or '',
            # Vitals
            "weight_kg": v.weight_kg or '',
            "height_cm": v.height_cm or '',
            "bmi": v.bmi_calculated or '',
            "bp_systolic": v.bp_systolic or '',
            "bp_diastolic": v.bp_diastolic or '',
            # Pregnancy
            "pdt_done": v.pdt_done or False,
            "pdt_result": v.pdt_result or '',
            # MEC
            "mec_conditions": v.mec_conditions or [],
            # Method
            "primary_method": v.primary_method or '',
            "method_visit_category": v.method_visit_category or '',
            "quantity_dispensed": v.quantity_dispensed or '',
            "dmpa_sc_mode": v.dmpa_sc_mode or '',
            "dmpa_take_home_doses": v.dmpa_sc_take_home_doses or '',
            "larc_removal_reason": v.larc_removal_reason or '',
            "return_date": v.return_date.strftime('%d/%m/%Y') if v.return_date else '',
            # HIV/STI
            "hiv_counselled": v.hiv_counselled or False,
            "hiv_tested": v.hiv_tested or False,
            "hiv_status": v.hiv_status or '',
            "tb_status": v.tb_status or '',
            "ipv_status": v.ipv_rc_status or '',
            "cervical_method": v.cervical_screening_method or '',
            "cervical_result": v.cervical_screening_result or '',
            # Condoms & Natural FP
            "condom_type": v.condoms_client_sex or '',
            "condom_qty": v.condoms_qty_dispensed or '',
            "natural_fp": v.natural_fp_counselled or False,
            "cycle_beads": v.cycle_beads_given or False,
            # PPFP & Referral
            "ppfp_timing": v.ppfp_timing or '',
            "referral_in": v.referral_in or '',
            "referral_out": v.referral_out or '',
            "remarks": v.remarks or '',
        })

    return {
        "period": f"{year or 'all'}-{str(month).zfill(2) if month else 'all'}",
        "total": len(result),
        "visits": result
    }